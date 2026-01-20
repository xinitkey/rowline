import os
import tempfile
import mimetypes
import subprocess
import asyncio
import concurrent.futures
from typing import Optional, List, Tuple
from pathlib import Path
import threading
import time

import pdfkit          # html -> pdf [web:1]
from docx2pdf import convert as docx2pdf_convert  # docx -> pdf [web:9]
import img2pdf         # images -> pdf [web:10]
from PIL import Image  # img2pdf зависит от Pillow [web:10]
from pypdf import PdfReader, PdfWriter


class UnsupportedFormat(Exception):
    pass


class ConversionProgress:
    """Track conversion progress for large files."""
    def __init__(self, total_steps: int = 1):
        self.total_steps = total_steps
        self.current_step = 0
        self.status = "Starting conversion..."
        self._lock = threading.Lock()

    def update(self, step: int, status: str):
        with self._lock:
            self.current_step = step
            self.status = status

    def get_progress(self) -> Tuple[int, str]:
        with self._lock:
            return self.current_step, self.status


# Global thread pool for CPU-bound operations - dynamically sized
import os
_cpu_executor = None

def get_cpu_executor():
    """Get or create CPU executor with optimal thread count."""
    global _cpu_executor
    if _cpu_executor is None:
        # Use system CPU count for optimal performance
        cpu_count = os.cpu_count() or 4
        # For CPU-bound operations like PDF conversion, use CPU count
        max_workers = max(cpu_count, 8)  # At least 8 threads
        _cpu_executor = concurrent.futures.ThreadPoolExecutor(
            max_workers=max_workers, 
            thread_name_prefix="pdf-convert"
        )
        print(f"[PDF] Initialized CPU thread pool with {max_workers} workers")
    return _cpu_executor


async def any_to_pdf_async(input_path: str, output_path: str | None = None, progress_callback: Optional[ConversionProgress] = None) -> str:
    """
    Async version of any_to_pdf with progress tracking and optimized performance.
    Convert file to PDF by extension with enhanced concurrency.
    """
    if output_path is None:
        base, _ = os.path.splitext(input_path)
        output_path = base + ".pdf"

    ext = os.path.splitext(input_path)[1].lower()

    if progress_callback:
        progress_callback.update(0, f"Analyzing file type: {ext}")

    if ext == ".pdf":
        # Already PDF - just copy asynchronously
        if input_path != output_path:
            await asyncio.to_thread(copy_file_sync, input_path, output_path)
        if progress_callback:
            progress_callback.update(1, "File is already PDF")
        return output_path

    # Initialize progress tracking
    if progress_callback:
        progress_callback.update(0, f"Converting {ext.upper()} to PDF")

    # Route to appropriate converter
    if ext in {".html", ".htm"}:
        await html_to_pdf_async(input_path, output_path, progress_callback)
    elif ext == ".xml":
        await xml_to_pdf_async(input_path, output_path, progress_callback)
    elif ext == ".docx":
        await docx_to_pdf_async(input_path, output_path, progress_callback)
    elif ext in {".xlsx", ".xls"}:
        await excel_to_pdf_async(input_path, output_path, progress_callback)
    elif ext in {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff"}:
        await image_to_pdf_async(input_path, output_path, progress_callback)
    elif ext in {".txt", ".py", ".log", ".md"}:
        await text_to_pdf_async(input_path, output_path, progress_callback)
    else:
        # Fallback by MIME type
        mime, _ = mimetypes.guess_type(input_path)
        if mime and mime.startswith("image/"):
            await image_to_pdf_async(input_path, output_path, progress_callback)
        else:
            raise UnsupportedFormat(f"Формат {ext} не поддерживается")

    if progress_callback:
        progress_callback.update(1, "Conversion completed")

    return output_path


def copy_file_sync(input_path: str, output_path: str) -> None:
    """Synchronous file copy for PDF files."""
    with open(input_path, "rb") as src, open(output_path, "wb") as dst:
        # Copy in chunks for large files
        while chunk := src.read(8192):
            dst.write(chunk)


async def html_to_pdf_async(input_path: str, output_path: str, progress: Optional[ConversionProgress] = None) -> None:
    """
    Async HTML to PDF conversion with optimized wkhtmltopdf usage.
    """
    if progress:
        progress.update(0.2, "Converting HTML to PDF")

    def _convert():
        return html_to_pdf_sync(input_path, output_path)

    await asyncio.get_event_loop().run_in_executor(get_cpu_executor(), _convert)

    if progress:
        progress.update(0.8, "HTML conversion completed")


async def xml_to_pdf_async(input_path: str, output_path: str, progress: Optional[ConversionProgress] = None) -> None:
    """
    Async XML to PDF conversion via HTML intermediate.
    """
    if progress:
        progress.update(0.3, "Converting XML to HTML")

    def _convert():
        return xml_to_pdf_sync(input_path, output_path)

    await asyncio.get_event_loop().run_in_executor(get_cpu_executor(), _convert)

    if progress:
        progress.update(0.9, "XML conversion completed")


async def docx_to_pdf_async(input_path: str, output_path: str, progress: Optional[ConversionProgress] = None) -> None:
    """
    Async DOCX to PDF conversion using LibreOffice.
    """
    if progress:
        progress.update(0.1, "Loading LibreOffice")

    def _convert():
        return docx_to_pdf_sync(input_path, output_path)

    await asyncio.get_event_loop().run_in_executor(get_cpu_executor(), _convert)

    if progress:
        progress.update(0.7, "DOCX conversion completed")


async def excel_to_pdf_async(input_path: str, output_path: str, progress: Optional[ConversionProgress] = None) -> None:
    """
    Async Excel to PDF conversion.
    """
    if progress:
        progress.update(0.2, "Converting Excel to PDF")

    def _convert():
        return excel_to_pdf_sync(input_path, output_path)

    await asyncio.get_event_loop().run_in_executor(get_cpu_executor(), _convert)

    if progress:
        progress.update(0.8, "Excel conversion completed")


async def image_to_pdf_async(input_path: str, output_path: str, progress: Optional[ConversionProgress] = None) -> None:
    """
    Async image to PDF conversion with optimization.
    """
    if progress:
        progress.update(0.4, "Processing image")

    def _convert():
        return image_to_pdf_sync(input_path, output_path)

    await asyncio.get_event_loop().run_in_executor(get_cpu_executor(), _convert)

    if progress:
        progress.update(0.9, "Image conversion completed")


async def text_to_pdf_async(input_path: str, output_path: str, progress: Optional[ConversionProgress] = None) -> None:
    """
    Async text to PDF conversion.
    """
    if progress:
        progress.update(0.5, "Converting text to PDF")

    def _convert():
        return text_to_pdf_sync(input_path, output_path)

    await asyncio.get_event_loop().run_in_executor(get_cpu_executor(), _convert)

    if progress:
        progress.update(0.9, "Text conversion completed")


# Synchronous versions for executor calls
def html_to_pdf_sync(input_path: str, output_path: str) -> None:
    """Synchronous HTML to PDF conversion."""
    html_to_pdf(input_path, output_path)


def xml_to_pdf_sync(input_path: str, output_path: str) -> None:
    """Synchronous XML to PDF conversion."""
    xml_to_pdf(input_path, output_path)


def docx_to_pdf_sync(input_path: str, output_path: str) -> None:
    """Synchronous DOCX to PDF conversion."""
    docx_to_pdf(input_path, output_path)


def excel_to_pdf_sync(input_path: str, output_path: str) -> None:
    """Synchronous Excel to PDF conversion."""
    excel_to_pdf(input_path, output_path)


def image_to_pdf_sync(input_path: str, output_path: str) -> None:
    """Synchronous image to PDF conversion."""
    image_to_pdf(input_path, output_path)


def text_to_pdf_sync(input_path: str, output_path: str) -> None:
    """Synchronous text to PDF conversion."""
    text_to_pdf(input_path, output_path)


# Legacy synchronous functions (keep for compatibility)
def any_to_pdf(input_path: str, output_path: str | None = None) -> str:
    """
    Synchronous version for backward compatibility.
    """
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        return loop.run_until_complete(any_to_pdf_async(input_path, output_path))
    finally:
        loop.close()


def html_to_pdf(input_path: str, output_path: str) -> None:
    """
    Convert HTML to PDF using pdfkit (wkhtmltopdf).
    Configured for local file conversion with proper options.
    """
    import os
    import tempfile
    import subprocess
    
    # Get the directory of the input file for resolving relative paths
    input_dir = os.path.dirname(os.path.abspath(input_path))
    
    # Set up environment for wkhtmltopdf
    env = os.environ.copy()
    env['XDG_RUNTIME_DIR'] = tempfile.gettempdir()
    env['QT_QPA_PLATFORM'] = 'offscreen'  # For headless operation
    
    # Try using wkhtmltopdf directly with proper environment
    wkhtmltopdf_path = None
    for path in ['wkhtmltopdf', '/usr/bin/wkhtmltopdf', '/usr/local/bin/wkhtmltopdf']:
        try:
            result = subprocess.run([path, '--version'], capture_output=True, env=env, timeout=10)
            if result.returncode == 0:
                wkhtmltopdf_path = path
                break
        except (subprocess.SubprocessError, FileNotFoundError):
            continue
    
    if wkhtmltopdf_path:
        # Use wkhtmltopdf directly
        cmd = [
            wkhtmltopdf_path,
            '--page-size', 'A4',
            '--margin-top', '0.75in',
            '--margin-right', '0.75in', 
            '--margin-bottom', '0.75in',
            '--margin-left', '0.75in',
            '--encoding', 'UTF-8',
            '--enable-local-file-access',
            '--disable-external-links',
            '--disable-internal-links',
            '--load-error-handling', 'ignore',
            '--load-media-error-handling', 'ignore',
            input_path,
            output_path
        ]
        
        try:
            result = subprocess.run(cmd, capture_output=True, text=True, encoding='utf-8', 
                                  env=env, timeout=60, check=True)
            print(f"[HTML] wkhtmltopdf conversion completed successfully")
            return
        except subprocess.CalledProcessError as e:
            print(f"[HTML] wkhtmltopdf failed: {e}")
        except subprocess.TimeoutExpired:
            print(f"[HTML] wkhtmltopdf timed out")
    
    # Fallback to pdfkit with minimal options
    print(f"[HTML] Using pdfkit fallback...")
    options = {
        'page-size': 'A4',
        'encoding': 'UTF-8',
        'disable-external-links': None,
        'disable-internal-links': None,
        'load-error-handling': 'ignore',
        'load-media-error-handling': 'ignore',
    }
    
    try:
        pdfkit.from_file(input_path, output_path, options=options)
    except Exception as e:
        print(f"[HTML] pdfkit fallback also failed: {e}")
        raise RuntimeError(f"HTML to PDF conversion failed: {e}")


def docx_to_pdf(input_path: str, output_path: str) -> None:
    """
    Convert DOCX to PDF using LibreOffice.
    LibreOffice provides good compatibility with Microsoft Office documents.
    """
    import shutil
    import os
    
    out_dir = os.path.dirname(os.path.abspath(output_path)) or "."
    
    # Find LibreOffice command
    libreoffice_cmd = None
    import platform
    if platform.system() == 'Linux':
        # Check common Linux installation paths
        possible_paths = [
            "/usr/bin/soffice",
            "/usr/bin/libreoffice",
            "/usr/local/bin/soffice",
            "/usr/local/bin/libreoffice",
            "/opt/libreoffice/program/soffice",
            "/snap/bin/soffice",  # Snap installation
            "/snap/bin/libreoffice",
        ]
        for path in possible_paths:
            if os.path.exists(path):
                libreoffice_cmd = path
                break
    elif platform.system() == 'Windows':
        # Check common Windows installation paths
        possible_paths = [
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
            r"C:\Program Files\LibreOffice 7\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice 7\program\soffice.exe",
        ]
        for path in possible_paths:
            if os.path.exists(path):
                libreoffice_cmd = path
                break
    # Also check PATH
    if not libreoffice_cmd:
        for cmd in ["libreoffice", "soffice", "libreoffice.exe", "soffice.exe"]:
            if shutil.which(cmd):
                libreoffice_cmd = cmd
                break
    
    if not libreoffice_cmd:
        raise UnsupportedFormat(
            "DOCX to PDF conversion requires LibreOffice. "
            "Install with: sudo apt install libreoffice"
        )
    
    # Log file size for debugging
    file_size = os.path.getsize(input_path)
    print(f"[DOCX] Converting file: {os.path.basename(input_path)} ({file_size} bytes)")
    
    # Check if file is extremely large (>100MB)
    if file_size > 100 * 1024 * 1024:
        print(f"[DOCX] Warning: Large file detected ({file_size} bytes). Conversion may take a long time.")
        try:
            import psutil
            import os
            process = psutil.Process(os.getpid())
            mem_before = process.memory_info().rss / 1024 / 1024  # MB
            print(f"[DOCX] Memory usage before conversion: {mem_before:.1f} MB")
        except ImportError:
            print("[DOCX] psutil not available for memory monitoring")
    
    try:
        result = subprocess.run(
            [
                libreoffice_cmd,
                "--headless",
                "--invisible",  # Run without GUI for better performance
                "--nocrashreport",  # Disable crash reporting
                "--nodefault",  # Don't load default modules
                "--convert-to", "pdf",
                "--outdir", out_dir,
                input_path
            ],
            capture_output=True,
            timeout=1800,  # Increased timeout for very large files (30 minutes)
            check=True,
            env={
                **os.environ,
                "XDG_CONFIG_HOME": tempfile.gettempdir(),
                "XDG_CACHE_HOME": tempfile.gettempdir()
            }
        )
        
        print(f"[DOCX] LibreOffice conversion completed successfully")
        
        # Log memory usage for large files
        if file_size > 100 * 1024 * 1024:
            try:
                mem_after = process.memory_info().rss / 1024 / 1024  # MB
                print(f"[DOCX] Memory usage after conversion: {mem_after:.1f} MB")
                print(f"[DOCX] Memory delta: {mem_after - mem_before:.1f} MB")
            except NameError:
                pass  # psutil not available
        
        # Log output safely (handle encoding issues)
        if result.stdout:
            safe_stdout = result.stdout.decode('utf-8', errors='replace')
            print(f"[DOCX] LibreOffice stdout: {safe_stdout}")
        if result.stderr:
            safe_stderr = result.stderr.decode('utf-8', errors='replace')
            print(f"[DOCX] LibreOffice stderr: {safe_stderr}")
        
        expected_pdf = os.path.join(out_dir, os.path.splitext(os.path.basename(input_path))[0] + ".pdf")
        if os.path.exists(expected_pdf):
            if expected_pdf != output_path:
                os.rename(expected_pdf, output_path)
            return
        
        raise RuntimeError("LibreOffice did not create PDF file")
        
    except subprocess.CalledProcessError as e:
        # Handle encoding issues in error output
        error_msg = str(e)
        if hasattr(e, 'stdout') and e.stdout:
            safe_stdout = e.stdout.decode('utf-8', errors='replace')
            error_msg += f" stdout: {safe_stdout}"
        if hasattr(e, 'stderr') and e.stderr:
            safe_stderr = e.stderr.decode('utf-8', errors='replace')
            error_msg += f" stderr: {safe_stderr}"
        raise RuntimeError(f"DOCX conversion failed: {error_msg}")
    except subprocess.TimeoutExpired:
        raise RuntimeError("DOCX conversion timed out")


def excel_to_pdf(input_path: str, output_path: str) -> None:
    """
    Convert Excel (XLSX/XLS) to PDF using LibreOffice with optimizations.
    Falls back to alternative methods for better performance.
    """
    import shutil
    import os
    
    out_dir = os.path.dirname(os.path.abspath(output_path)) or "."
    
    # Get CPU count for optimizations
    cpu_count = os.cpu_count() or 4
    
    # Find LibreOffice command
    libreoffice_cmd = None
    import platform
    if platform.system() == 'Linux':
        # Check common Linux installation paths
        possible_paths = [
            "/usr/bin/soffice",
            "/usr/bin/libreoffice",
            "/usr/local/bin/soffice",
            "/usr/local/bin/libreoffice",
            "/opt/libreoffice/program/soffice",
            "/snap/bin/soffice",  # Snap installation
            "/snap/bin/libreoffice",
        ]
        for path in possible_paths:
            if os.path.exists(path):
                libreoffice_cmd = path
                break
    elif platform.system() == 'Windows':
        # Check common Windows installation paths
        possible_paths = [
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
            r"C:\Program Files\LibreOffice 7\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice 7\program\soffice.exe",
        ]
        for path in possible_paths:
            if os.path.exists(path):
                libreoffice_cmd = path
                break
    # Also check PATH
    if not libreoffice_cmd:
        for cmd in ["libreoffice", "soffice", "libreoffice.exe", "soffice.exe"]:
            if shutil.which(cmd):
                libreoffice_cmd = cmd
                break
    
    if not libreoffice_cmd:
        # Try alternative conversion method using openpyxl + reportlab
        try:
            return excel_to_pdf_alternative(input_path, output_path)
        except Exception as e:
            raise UnsupportedFormat(
                f"Excel to PDF conversion requires LibreOffice or alternative libraries. "
                f"LibreOffice not found, alternative method failed: {e}. "
                f"Install with: sudo apt install libreoffice"
            )
    
    # Log file size for debugging
    file_size = os.path.getsize(input_path)
    print(f"[Excel] Converting file: {os.path.basename(input_path)} ({file_size} bytes)")
    
    # Check if file is extremely large (>100MB) - use special handling
    if file_size > 100 * 1024 * 1024:
        print(f"[Excel] Warning: Large file detected ({file_size} bytes). Using optimized conversion.")
        return excel_to_pdf_large_file(input_path, output_path, libreoffice_cmd, cpu_count)
    
    try:
        # Optimize LibreOffice for performance
        env_vars = {
            **os.environ,
            "XDG_CONFIG_HOME": tempfile.gettempdir(),
            "XDG_CACHE_HOME": tempfile.gettempdir(),
            # Performance optimizations
            "SAL_USE_VCLPLUGIN": "svp",  # Use virtual device for headless
            "OOO_EXIT_POST_STARTUP": "1",  # Exit after startup
            # Memory and threading optimizations
            "SAL_MAXTHREADS": str(min(cpu_count * 2, 16)),  # Limit threads but allow parallelism
            "PARALLELISM_MAX": str(min(cpu_count, 8)),  # Allow parallel processing
        }
        
        result = subprocess.run(
            [
                libreoffice_cmd,
                "--headless",
                "--invisible",  # Run without GUI for better performance
                "--nocrashreport",  # Disable crash reporting
                "--nodefault",  # Don't load default modules
                "--convert-to", "pdf:calc_pdf_Export",
                "--outdir", out_dir,
                # Performance flags
                "--norestore",  # Don't restore previous session
                "--nolockcheck",  # Skip lock file checks
                input_path
            ],
            capture_output=True,
            timeout=1800,  # Increased timeout for very large files (30 minutes)
            check=True,
            env=env_vars
        )
        
        print(f"[Excel] LibreOffice conversion completed successfully")
        
        # Log output safely (handle encoding issues)
        if result.stdout:
            safe_stdout = result.stdout.decode('utf-8', errors='replace')
            print(f"[Excel] LibreOffice stdout: {safe_stdout}")
        if result.stderr:
            safe_stderr = result.stderr.decode('utf-8', errors='replace')
            print(f"[Excel] LibreOffice stderr: {safe_stderr}")
        
        expected_pdf = os.path.join(out_dir, os.path.splitext(os.path.basename(input_path))[0] + ".pdf")
        if os.path.exists(expected_pdf):
            if expected_pdf != output_path:
                os.rename(expected_pdf, output_path)
            return
        
        raise RuntimeError("LibreOffice did not create PDF file")
        
    except subprocess.CalledProcessError as e:
        # Handle encoding issues in error output
        error_msg = str(e)
        if hasattr(e, 'stdout') and e.stdout:
            safe_stdout = e.stdout.decode('utf-8', errors='replace')
            error_msg += f" stdout: {safe_stdout}"
        if hasattr(e, 'stderr') and e.stderr:
            safe_stderr = e.stderr.decode('utf-8', errors='replace')
            error_msg += f" stderr: {safe_stderr}"
        raise RuntimeError(f"Excel conversion failed: {error_msg}")
    except subprocess.TimeoutExpired:
        raise RuntimeError("Excel conversion timed out")


def image_to_pdf(input_path: str, output_path: str) -> None:
    # img2pdf.convert возвращает bytes PDF [web:10]
    with open(output_path, "wb") as f_out:
        f_out.write(img2pdf.convert(input_path))


def xml_to_pdf(input_path: str, output_path: str) -> None:
    """
    Convert XML to PDF with formatting.
    """
    import html as html_module
    
    with open(input_path, "r", encoding="utf-8", errors="ignore") as f:
        xml_content = f.read()
    
    # Escape HTML entities and wrap in pre tag for formatting
    escaped_xml = html_module.escape(xml_content)
    
    html = f"""
    <html>
      <head>
        <meta charset="utf-8">
        <style>
          body {{
            font-family: 'Consolas', 'Monaco', monospace;
            font-size: 10pt;
            line-height: 1.4;
            padding: 20px;
            background: #f8f9fa;
          }}
          pre {{
            white-space: pre-wrap;
            word-wrap: break-word;
            background: white;
            padding: 15px;
            border: 1px solid #ddd;
            border-radius: 4px;
          }}
        </style>
      </head>
      <body><pre>{escaped_xml}</pre></body>
    </html>
    """
    
    with tempfile.NamedTemporaryFile("w", suffix=".html", delete=False, encoding="utf-8") as tmp_html:
        tmp_html.write(html)
        tmp_html_path = tmp_html.name
    
    try:
        pdfkit.from_file(tmp_html_path, output_path)
    finally:
        os.remove(tmp_html_path)


def text_to_pdf(input_path: str, output_path: str) -> None:
    """
    Simple: wrap text in HTML and render to PDF via wkhtmltopdf.
    Can be replaced with fpdf/reportlab if you don't want to depend on wkhtmltopdf.
    """
    with open(input_path, "r", encoding="utf-8", errors="ignore") as f:
        text = f.read()

    html = f"""
    <html>
      <head>
        <meta charset="utf-8">
        <style>
          body {{
            font-family: monospace;
            white-space: pre-wrap;
          }}
        </style>
      </head>
      <body>{text}</body>
    </html>
    """

    with tempfile.NamedTemporaryFile("w", suffix=".html", delete=False, encoding="utf-8") as tmp_html:
        tmp_html.write(html)
        tmp_html_path = tmp_html.name

    try:
        pdfkit.from_file(tmp_html_path, output_path)
    finally:
        os.remove(tmp_html_path)


def split_pdf(input_path: str, output_dir: str, pages: list[int] | None = None) -> list[str]:
    """
    Split PDF into multiple files with optimized parallel processing.

    Args:
        input_path: Path to input PDF
        output_dir: Directory to save split files
        pages: List of page numbers to extract (1-based), if None - split all pages individually

    Returns:
        List of output file paths
    """
    reader = PdfReader(input_path)
    total_pages = len(reader.pages)

    if pages is None:
        # Split each page into separate file - use parallel processing for large PDFs
        if total_pages <= 10:
            return split_pdf_sequential(reader, output_dir, total_pages)
        else:
            return split_pdf_parallel(reader, output_dir, total_pages)
    else:
        # Extract specific pages into one file
        writer = PdfWriter()
        for page_num in pages:
            if 1 <= page_num <= total_pages:
                writer.add_page(reader.pages[page_num - 1])

        output_path = os.path.join(output_dir, "extracted_pages.pdf")
        with open(output_path, "wb") as f:
            writer.write(f)
        return [output_path]


def split_pdf_sequential(reader: PdfReader, output_dir: str, total_pages: int) -> list[str]:
    """Sequential PDF splitting for small number of pages."""
    output_files = []
    for i in range(total_pages):
        writer = PdfWriter()
        writer.add_page(reader.pages[i])

        output_path = os.path.join(output_dir, f"page_{i+1:03d}.pdf")
        with open(output_path, "wb") as f:
            writer.write(f)
        output_files.append(output_path)
    return output_files


def split_pdf_parallel(reader: PdfReader, output_dir: str, total_pages: int) -> list[str]:
    """
    Parallel PDF splitting for better performance with many pages.
    Uses thread pool to write multiple PDF files concurrently.
    """
    import concurrent.futures

    def write_single_page(page_idx: int) -> str:
        """Write a single page to PDF file."""
        writer = PdfWriter()
        writer.add_page(reader.pages[page_idx])

        output_path = os.path.join(output_dir, f"page_{page_idx+1:03d}.pdf")
        with open(output_path, "wb") as f:
            writer.write(f)
        return output_path

    # Write all pages in parallel
    with concurrent.futures.ThreadPoolExecutor(max_workers=min(total_pages, 8)) as executor:
        # Submit all write tasks
        futures = [executor.submit(write_single_page, i) for i in range(total_pages)]

        # Collect results
        output_files = []
        for future in concurrent.futures.as_completed(futures):
            try:
                output_files.append(future.result())
            except Exception as e:
                print(f"[SPLIT] Error writing page: {e}")
                raise

    return sorted(output_files)  # Return in page order
def merge_pdf(input_paths: list[str], output_path: str) -> str:
    """
    Merge multiple PDF files into one with optimized parallel processing.

    Args:
        input_paths: List of PDF file paths to merge
        output_path: Path for merged output file

    Returns:
        Output file path
    """
    if len(input_paths) <= 3:
        # For small number of files, use simple sequential merge
        return merge_pdf_sequential(input_paths, output_path)
    else:
        # For larger number of files, use parallel processing
        return merge_pdf_parallel(input_paths, output_path)


def merge_pdf_sequential(input_paths: list[str], output_path: str) -> str:
    """Sequential PDF merging for small number of files."""
    writer = PdfWriter()

    for pdf_path in input_paths:
        reader = PdfReader(pdf_path)
        for page in reader.pages:
            writer.add_page(page)

    with open(output_path, "wb") as f:
        writer.write(f)

    return output_path


def merge_pdf_parallel(input_paths: list[str], output_path: str) -> str:
    """
    Parallel PDF merging for better performance with many files.
    Uses thread pool to read multiple PDF files concurrently.
    """
    import concurrent.futures

    writer = PdfWriter()

    def read_pdf_pages(pdf_path: str) -> list:
        """Read all pages from a PDF file."""
        reader = PdfReader(pdf_path)
        return list(reader.pages)

    # Read all PDF files in parallel
    with concurrent.futures.ThreadPoolExecutor(max_workers=min(len(input_paths), 8)) as executor:
        # Submit all read tasks
        future_to_path = {executor.submit(read_pdf_pages, path): path for path in input_paths}

        # Collect results as they complete
        for future in concurrent.futures.as_completed(future_to_path):
            try:
                pages = future.result()
                # Add all pages from this PDF to the writer
                for page in pages:
                    writer.add_page(page)
            except Exception as e:
                print(f"[MERGE] Error reading PDF {future_to_path[future]}: {e}")
                raise

    # Write the merged PDF
    with open(output_path, "wb") as f:
        writer.write(f)

    return output_path


def excel_to_pdf_large_file(input_path: str, output_path: str, libreoffice_cmd: str, cpu_count: int) -> None:
    """
    Optimized conversion for large Excel files using LibreOffice with performance tuning.
    """
    import os
    import tempfile
    
    out_dir = os.path.dirname(os.path.abspath(output_path)) or "."
    
    # Create temporary working directory to avoid conflicts
    with tempfile.TemporaryDirectory() as temp_dir:
        # Copy file to temp directory for processing
        temp_input = os.path.join(temp_dir, os.path.basename(input_path))
        import shutil
        shutil.copy2(input_path, temp_input)
        
        try:
            # Optimize LibreOffice for performance with large files
            env_vars = {
                **os.environ,
                "XDG_CONFIG_HOME": temp_dir,
                "XDG_CACHE_HOME": temp_dir,
                # Performance optimizations for large files
                "SAL_USE_VCLPLUGIN": "svp",  # Use virtual device for headless
                "OOO_EXIT_POST_STARTUP": "1",  # Exit after startup
                # Memory and threading optimizations
                "SAL_MAXTHREADS": str(min(cpu_count * 2, 32)),  # More threads for large files
                "PARALLELISM_MAX": str(min(cpu_count, 16)),  # Allow parallel processing
                # Disable unnecessary features for speed
                "DISABLE_OPENCL": "1",  # Disable OpenCL to avoid GPU issues
                "SAL_DISABLE_FLOATGRABBER": "1",  # Disable floating point optimizations
            }
            
            result = subprocess.run(
                [
                    libreoffice_cmd,
                    "--headless",
                    "--invisible",
                    "--nocrashreport",
                    "--nodefault",
                    "--convert-to", "pdf:calc_pdf_Export",
                    "--outdir", temp_dir,
                    # Performance flags for large files
                    "--norestore",
                    "--nolockcheck",
                    "--quickstart=no",  # Disable quickstart for better performance
                    temp_input
                ],
                capture_output=True,
                timeout=3600,  # 1 hour for very large files
                check=True,
                env=env_vars
            )
            
            print(f"[Excel-Large] LibreOffice conversion completed successfully")
            
            # Find and move the output file
            expected_pdf = os.path.join(temp_dir, os.path.splitext(os.path.basename(input_path))[0] + ".pdf")
            if os.path.exists(expected_pdf):
                shutil.move(expected_pdf, output_path)
                return
            
            raise RuntimeError("LibreOffice did not create PDF file for large Excel")
            
        except subprocess.CalledProcessError as e:
            error_msg = f"LibreOffice conversion failed: {e}"
            if hasattr(e, 'stdout') and e.stdout:
                safe_stdout = e.stdout.decode('utf-8', errors='replace')
                error_msg += f" stdout: {safe_stdout}"
            if hasattr(e, 'stderr') and e.stderr:
                safe_stderr = e.stderr.decode('utf-8', errors='replace')
                error_msg += f" stderr: {safe_stderr}"
            raise RuntimeError(error_msg)


def excel_to_pdf_alternative(input_path: str, output_path: str) -> None:
    """
    Alternative Excel to PDF conversion using Python libraries (fallback method).
    Less accurate but faster for simple spreadsheets.
    """
    try:
        from openpyxl import load_workbook
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib import colors
        import os
        
        print(f"[Excel-Alt] Using alternative conversion method for {os.path.basename(input_path)}")
        
        # Load workbook
        wb = load_workbook(input_path, read_only=True, data_only=True)
        
        # Create PDF
        doc = SimpleDocTemplate(output_path, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Add sheet title
            title = Paragraph(f"Sheet: {sheet_name}", styles['Heading1'])
            elements.append(title)
            
            # Extract data
            data = []
            for row in ws.iter_rows(values_only=True):
                # Convert all values to strings and handle None
                row_data = [str(cell) if cell is not None else "" for cell in row]
                if any(row_data):  # Skip empty rows
                    data.append(row_data)
            
            if data:
                # Create table
                table = Table(data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 14),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(table)
        
        doc.build(elements)
        print(f"[Excel-Alt] Alternative conversion completed: {os.path.basename(output_path)}")
        
    except ImportError as e:
        raise UnsupportedFormat(f"Alternative Excel conversion requires openpyxl and reportlab: {e}")
    except Exception as e:
        raise RuntimeError(f"Alternative Excel conversion failed: {e}")
