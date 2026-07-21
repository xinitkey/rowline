"""
XLSX file reader.
"""

from pathlib import Path
from typing import Optional, Iterator
from dataclasses import dataclass

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class SheetData:
    """Excel sheet data."""
    name: str
    headers: list[str]
    rows: list[list]


class XlsxReader:
    """XLSX file reader."""

    def __init__(self, file_path: str | Path):
        """
        Args:
            file_path: Path to XLSX file.
        """
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"File not found: {self.file_path}")
        if not self.file_path.suffix.lower() == ".xlsx":
            raise ValueError(f"Expected .xlsx file, got: {self.file_path.suffix}")

        self._workbook = load_workbook(self.file_path, read_only=True, data_only=True)

    @property
    def sheet_names(self) -> list[str]:
        """Return list of sheet names."""
        return self._workbook.sheetnames

    def read_sheet(
        self,
        sheet_name: Optional[str] = None,
        header_row: int = 1,
        start_row: int = 2,
        num_columns: int = 9
    ) -> SheetData:
        """
        Read data from the specified sheet.

        Args:
            sheet_name: Sheet name (default: active sheet).
            header_row: Header row number (1-based).
            start_row: Data start row number (1-based).
            num_columns: Number of columns to read.

        Returns:
            SheetData with headers and data.
        """
        if sheet_name:
            sheet: Worksheet = self._workbook[sheet_name]
        else:
            sheet = self._workbook.active

        # Read headers
        headers = []
        for cell in list(sheet[header_row])[:num_columns]:
            value = cell.value
            if value is not None:
                headers.append(self._sanitize_header(str(value)))
            else:
                headers.append(f"column_{len(headers) + 1}")

        # Read data (keep all rows, including empty ones)
        rows = []
        for row in sheet.iter_rows(min_row=start_row, max_col=num_columns):
            row_data = [cell.value for cell in row]
            rows.append(row_data)

        return SheetData(
            name=sheet.title,
            headers=headers,
            rows=rows
        )

    def read_all_sheets(
        self, 
        header_row: int = 1, 
        start_row: int = 2,
        num_columns: int = 9
    ) -> Iterator[SheetData]:
        """
        Read data from all sheets.

        Args:
            header_row: Header row number.
            start_row: Data start row number.
            num_columns: Number of columns to read.

        Yields:
            SheetData for each sheet.
        """
        for sheet_name in self.sheet_names:
            yield self.read_sheet(sheet_name, header_row, start_row, num_columns)

    @staticmethod
    def _sanitize_header(header: str) -> str:
        """
        Sanitize header for use in XML.

        Args:
            header: Raw header string.

        Returns:
            Sanitized header.
        """
        result = ""
        for char in header.strip():
            if char.isalnum() or char == "_":
                result += char
            elif char in " -":
                result += "_"

        if result and result[0].isdigit():
            result = "_" + result

        return result or "field"

    def close(self):
        """Close the workbook."""
        self._workbook.close()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
