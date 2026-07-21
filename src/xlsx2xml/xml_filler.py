from pathlib import Path
from typing import Optional, Any
from datetime import datetime

from lxml import etree
from openpyxl import load_workbook


class XmlFiller:
    def __init__(self, template_path: str | Path):
        self.template_path = Path(template_path)
        if not self.template_path.exists():
            raise FileNotFoundError(f"Template not found: {self.template_path}")

        parser = etree.XMLParser(remove_blank_text=False)
        self._template_root = etree.parse(str(self.template_path), parser).getroot()
        self._template_bytes = etree.tostring(self._template_root, encoding='unicode')

        self._row_index: dict[str, etree._Element] = {}
        for row in self._template_root.findall(".//Row"):
            code = row.get("code", "")
            if code:
                self._row_index[code] = row

    def _copy_template(self) -> tuple[etree._Element, dict[str, etree._Element]]:
        root = etree.fromstring(self._template_bytes)
        idx = {}
        for row in root.iter("Row"):
            code = row.get("code")
            if code:
                idx[code] = row
        return root, idx

    def fill_from_xlsx(
        self,
        xlsx_path: str | Path,
        output_path: str | Path,
        sheet_name: Optional[str] = None,
        code_column: int = 6,
        data_start_column: int = 7,
        data_end_column: int = 12,
        start_row: int = 9,
    ) -> Path:
        xlsx_path = Path(xlsx_path)
        output_path = Path(output_path)
        if not xlsx_path.exists():
            raise FileNotFoundError(f"XLSX not found: {xlsx_path}")

        filled_root, row_index = self._copy_template()
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        sheet = wb[sheet_name] if sheet_name else wb.active

        for row in sheet.iter_rows(min_row=start_row, min_col=1, max_col=data_end_column):
            code_val = row[code_column - 1].value
            if code_val is None:
                continue
            code_str = self._normalize_code(code_val)
            if not code_str or code_str not in row_index:
                continue

            values = [row[i].value if i < len(row) else None for i in range(data_start_column - 1, data_end_column)]
            self._update_row_values(row_index[code_str], values)

        wb.close()
        output_path.parent.mkdir(parents=True, exist_ok=True)
        etree.ElementTree(filled_root).write(str(output_path), encoding="UTF-8", xml_declaration=True, pretty_print=True)
        return output_path

    def _normalize_code(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return f"{value.day}.{value.month}"
        if isinstance(value, float):
            return str(int(value)) if value.is_integer() else str(value)
        return str(value).strip()

    def _update_row_values(self, xml_row: etree._Element, values: list) -> bool:
        col_idx = {col.get("code"): col for col in xml_row.iter("Col")}
        updated = False
        for i, code in enumerate(("4", "5", "6", "7", "8", "9")):
            if i >= len(values):
                break
            col = col_idx.get(code)
            if col is None or col.get("isNumerical") != "true":
                continue
            val = values[i]
            if val is not None:
                new = self._format_value(val)
                if new and col.text != new:
                    col.text = new
                    updated = True
        return updated

    def _format_value(self, value: Any) -> Optional[str]:
        if value is None:
            return None
        if isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            s = str(int(value)) if isinstance(value, float) and value.is_integer() else str(round(value, 2))
            return s
        s = str(value).strip()
        if not s:
            return None
        try:
            n = float(s.replace(",", "."))
            return str(int(n)) if n.is_integer() else str(round(n, 2))
        except ValueError:
            return None

    def fill_all_sheets(
        self,
        xlsx_path: str | Path,
        output_dir: str | Path,
        code_column: int = 6,
        data_start_column: int = 7,
        data_end_column: int = 12,
        start_row: int = 9,
    ) -> list[Path]:
        xlsx_path = Path(xlsx_path)
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        results = []
        for sheet_name in wb.sheetnames:
            safe = "".join(c if c.isalnum() or c in "._-" else "_" for c in sheet_name)
            out = output_dir / f"{safe}.xml"
            try:
                root, idx = self._copy_template()
                sheet = wb[sheet_name]
                for row in sheet.iter_rows(min_row=start_row, min_col=1, max_col=data_end_column):
                    code_val = row[code_column - 1].value
                    if code_val is None:
                        continue
                    code_str = self._normalize_code(code_val)
                    if not code_str or code_str not in idx:
                        continue
                    values = [row[i].value if i < len(row) else None for i in range(data_start_column - 1, data_end_column)]
                    self._update_row_values(idx[code_str], values)
                etree.ElementTree(root).write(str(out), encoding="UTF-8", xml_declaration=True, pretty_print=True)
                results.append(out)
            except Exception:
                pass
        wb.close()
        return results
