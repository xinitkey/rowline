"""
Модуль для чтения XLSX файлов.
"""

from pathlib import Path
from typing import Optional, Iterator
from dataclasses import dataclass

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class SheetData:
    """Данные листа Excel."""
    name: str
    headers: list[str]
    rows: list[list]


class XlsxReader:
    """Класс для чтения XLSX файлов."""

    def __init__(self, file_path: str | Path):
        """
        Инициализация читателя XLSX.

        Args:
            file_path: Путь к XLSX файлу.
        """
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Файл не найден: {self.file_path}")
        if not self.file_path.suffix.lower() == ".xlsx":
            raise ValueError(f"Ожидается файл .xlsx, получен: {self.file_path.suffix}")

        self._workbook = load_workbook(self.file_path, read_only=True, data_only=True)

    @property
    def sheet_names(self) -> list[str]:
        """Возвращает список имен листов."""
        return self._workbook.sheetnames

    def read_sheet(
        self,
        sheet_name: Optional[str] = None,
        header_row: int = 1,
        start_row: int = 2,
        num_columns: int = 9
    ) -> SheetData:
        """
        Читает данные с указанного листа.

        Args:
            sheet_name: Имя листа (по умолчанию - активный лист).
            header_row: Номер строки с заголовками (1-based).
            start_row: Номер строки начала данных (1-based).
            num_columns: Количество колонок для чтения.

        Returns:
            SheetData с заголовками и данными.
        """
        if sheet_name:
            sheet: Worksheet = self._workbook[sheet_name]
        else:
            sheet = self._workbook.active

        # Читаем заголовки
        headers = []
        for cell in list(sheet[header_row])[:num_columns]:
            value = cell.value
            if value is not None:
                headers.append(self._sanitize_header(str(value)))
            else:
                headers.append(f"column_{len(headers) + 1}")

        # Читаем данные (сохраняем все строки, включая пустые)
        rows = []
        for row in sheet.iter_rows(min_row=start_row, max_col=num_columns):
            row_data = [cell.value for cell in row]
            rows.append(row_data)

        return SheetData(
            name=sheet.title,
            headers=headers,
            rows=rows
        )

    def read_all_sheets(
        self, 
        header_row: int = 1, 
        start_row: int = 2,
        num_columns: int = 9
    ) -> Iterator[SheetData]:
        """
        Читает данные со всех листов.

        Args:
            header_row: Номер строки с заголовками.
            start_row: Номер строки начала данных.
            num_columns: Количество колонок для чтения.

        Yields:
            SheetData для каждого листа.
        """
        for sheet_name in self.sheet_names:
            yield self.read_sheet(sheet_name, header_row, start_row, num_columns)

    @staticmethod
    def _sanitize_header(header: str) -> str:
        """
        Очищает заголовок для использования в XML.

        Args:
            header: Исходный заголовок.

        Returns:
            Очищенный заголовок.
        """
        result = ""
        for char in header.strip():
            if char.isalnum() or char == "_":
                result += char
            elif char in " -":
                result += "_"

        if result and result[0].isdigit():
            result = "_" + result

        return result or "field"

    def close(self):
        """Закрывает рабочую книгу."""
        self._workbook.close()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
